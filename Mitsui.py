# -*- coding: utf-8 -*-
__author__ = 'newdefence@163.com'
__date__ = '2022/08/11 15:23' 

# import sys
# reload(sys)
# sys.setdefaultencoding('utf8')

import logging
from decimal import Decimal

from openpyxl import load_workbook
from openpyxl.cell.read_only import EmptyCell

logger = logging.getLogger()

# TODO: excel 文件需要统一表头(第二行表头文字一定要一致，否则没法处理，或者需要根据一个文件一种处理方式)


def read_invoice(workbook):
    sheet = workbook.worksheets[0]
    # 发票号，PO号，物料号，数量，单价，合计，总数量，总合计，总件数，总毛重，总净重
    # invoice, po, cpn = None, None, None
    # qty, unit_value, total_value, total_qty, total_invoice_value, total_pkg = None, None, None, None, None, None
    # total_gross_weight, total_net_weight = None, None
    columns = None  # { '发票号': 1, '物料号': 3, ... }
    # { 'invoice_no1': { 'sum': { '总合计': 1, '总毛重': 2, ... }, 'details': [{ 'PO号': 'xxx', '物料号': '', ... }, ...] } }
    invoices, all_sum = {}, None
    for ri, row in enumerate(sheet.iter_rows()):
        if not ri:  # 首行表头忽略
            continue
        if ri == 1:  # 第二行根据中文提取字段
            columns = dict((cell.value, cell.column - 1) for cell in row if not isinstance(cell, EmptyCell))
        else:
            # 汇总行只有一行，如果有多个，则出错
            cell_a_n = row[0]
            if isinstance(cell_a_n, EmptyCell):
                all_sum = {
                    '总数量': row[columns['总数量']].value,
                    '总合计': row[columns['总合计']].value,
                    '总毛重': row[columns['总毛重']].value,
                    # NOTE: 为什么总件数没有呢？
                }
            else:
                # 只处理所需的字段，由于涉及到空单元格的问题，因此需要逐个字段处理，如果为空则按0处理
                fields = ('发票号', 'PO号（明细）', '物料号', '数量', '单价', '合计', '总数量', '总合计', '总件数', '总毛重')
                detail = {}
                for key in fields:
                    cell = row[columns[key]]
                    detail[key] = 0 if isinstance(cell, EmptyCell) else cell.value
                invoice = invoices.setdefault(cell_a_n.value, {'details': [], 'sum': detail})
                invoice['details'].append(detail)
                # NOTE: 浮点数计算问题不可忽略
                if detail['合计'] != detail['数量'] * detail['单价']:
                    print(detail)
                    raise Exception('TODO')
    # 每一个发票号明细总数量全部相等，并等于该发票号所有数量合计
    for invoice in invoices.values():
        sum1, details = invoice['sum'], invoice['details']
        total_qty, total_invoice_value, total_gross_weight, total_pkgs = sum1['总数量'], sum1['总合计'], sum1['总毛重'], sum1['总件数']
        if total_qty != sum(r['数量'] for r in details):
            raise Exception('总数量与数量总和不符')
        if not all(r['总数量'] == total_qty for r in details):
            raise Exception('总数量错误')
        if total_invoice_value != sum(r['合计'] for r in details):
            raise Exception('总合计与合计总和不符')
        if not all(r['总合计'] == total_invoice_value for r in details):
            raise Exception('总合计错误')
        # 毛重没法比对，只能和自己比对
        if not all(r['总毛重'] == total_gross_weight for r in details):
            raise Exception('总毛重错误')
        if not all(r['总件数'] == total_pkgs for r in details):
            raise Exception('总件数错误')



def read_packing(workbook):
    # 发票号，PO号，物料号，数量，单价，合计，总数量，总合计，总件数，总毛重，总净重
    print('todo: packing')


def read_air(workbook):
    print('todo: air')


def check(proforma_invoice, packing_list, air_warbill):
    if proforma_invoice is None:
        logger.warning('无发票文件')
        return
    invoice = load_workbook(proforma_invoice, read_only=True)
    packing = load_workbook(packing_list, read_only=True) if packing_list else None
    air = load_workbook(air_warbill, read_only=True) if air_warbill else None

    # excel 文件其实没必要写三个，一个文件三个 sheet 即可
    read_invoice(invoice)
    if packing:
        read_packing(packing)
    if air:
        read_air(air)

