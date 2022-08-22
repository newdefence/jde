# -*- coding: utf-8 -*-
__author__ = 'newdefence@163.com'
__date__ = '2022/08/16 16:33' 

from decimal import Decimal
import logging
import re

from openpyxl import load_workbook

import checker
import reader

'''
立寰：concord
只有物流园和苏州有Excel

物流园
1.（所有供应商统一逻辑）发票跟箱单检验逻辑：相同的发票号单个料号总数量校验，单个发票号总数量校验，所有发票号总数量校验
箱单上总毛重大于总净重

    Excel校验：

    发票跟excel检验
    Po号、物料号进行数量、单价和合计的检验（excel单价进行保留四位小数，四舍五入后跟发票单价进行检验）（后续关务系统，单价以excel中超过4位的为准）
'''


logger = logging.getLogger()
reKG = re.compile(r'\(([\d,.]+)\s+KG\)') # (8.760 KG), 10.211LB (1.234 KG)
reEA = re.compile(r'([\d,.]+)\s+EA') # 60,000  EA
reOF = re.compile(r'\d+\s+of\s+(\d+)') # 01 of 02
re00 = re.compile(r'^0+') # 箱单发票号前置的00需要去掉


def Decimal2(rs):
    return Decimal(rs[0].replace(',', ''))


def read_invoice(sheet):
    """文件自查"""
    # 发票号，地址，原产国，PO号，物料号，数量，单价，合计，总数量，总合计，总件数，总毛重，总净重
    # { '发票号': 1, '物料号': 3, ... }
    # { 'invoice_no': { 'sum': { '总合计': 1, '总毛重': 2, ... }, 'details': [{ 'PO号': 'xxx', '物料号': '', ... }, ...] } }

    # NOTE: 立寰发票文件没有汇总行
    invoices, columns = reader.read_sheet12(sheet, ('发票号', 'po号', '物料号'),
                                    ('数量', '单价', '合计', '总数量', '总净重', '总合计', '总件数', '总毛重'))
    all_invoices = invoices.values()
    for invoice in all_invoices:
        checker.check_qty12(invoice)
        checker.check_invoice_value1(invoice)
        # 毛重没法比对，只能和自己比对
        checker.check_gross_weight1(invoice)
        checker.check_net_weight1(invoice)
        checker.check_piece12(invoice)
    # NOTE: 没有汇总核对需求
    all_sum ={
        '总数量': sum(v['sum']['总数量'] for v in all_invoices),
        '总合计': sum(v['sum']['总合计'] for v in all_invoices),
        '总毛重': sum(v['sum']['总毛重'] for v in all_invoices),
        '总净重': sum(v['sum']['总净重'] for v in all_invoices),
        '总件数': sum(v['sum']['总件数'] for v in all_invoices),
    }
    logger.info('发票文件核对完成')
    return columns, invoices, all_sum


def read_packing(sheet):
    packings, columns = reader.read_sheet12(sheet, ('发票号', 'po号', '物料号', '数量', '净重', '毛重', '总数量', '总净重', '总毛重', '总件数', '总托盘数'), ('总箱数',))
    # 每一个发票号明细总数量，总净重，总毛重全部相等，总件数全部相同，并等于该发票号所有数量合计
    all_pkgs = packings.values()
    for invoice in all_pkgs:
        sum1, details = invoice['sum'], invoice['details']
        # TODO NOTE: 修正净重，毛重：只有 发票号，物料，总件数 分组的第一个才有意义，其他的归值为0
        eraser = set()
        for d in details:
            if d['总件数'] in eraser:
                # TODO: 记录擦除，还是等待Excel修改？
                d['净重'] = 0
                d['毛重'] = 0
            else:
                eraser.add(d['总件数'])
            d['数量'] = Decimal2(reEA.findall(d['数量'])) # '60,000 EA' -> 60000
            d['总数量'] = Decimal2(reEA.findall(d['总数量'])) # '60,000 EA' -> 60000
            # 正常情况下毛重净重需要处理千位分隔符问题；此处处理一下以防万一
            d['净重'] = Decimal2(reKG.findall(d['净重'])) if d['净重'] else 0 # '(5.710 KG)' -> 5.710
            d['毛重'] = Decimal2(reKG.findall(d['毛重'])) if d['毛重'] else 0
            d['总净重'] = Decimal2(reKG.findall(d['总净重'])) if d['总净重'] else 0 # '8.952 LB (4.061 KG)' -> 4.061
            d['总毛重'] = Decimal2(reKG.findall(d['总毛重'])) if d['总毛重'] else 0
            # d['_总托盘数'] = d['总托盘数']
            d['_总件数'] = Decimal(reOF.findall(d['总件数'])[0])
        checker.check_qty12(invoice)
        checker.check_net_weight2(invoice)
        checker.check_gross_weight2(invoice)
        # checker.check_piece12(invoice)
        # 总件数需要特殊处理
        total_pieces = invoice['sum']['_总件数']
        tuple(d['errors'].add('总件数: %s %s' % (d['_总件数'], total_pieces)) for d in invoice['details'] if (d['_总件数'] != total_pieces))
    logger.info('箱单文件核对完成')
    all_sum = {
        '总数量': sum(v['sum']['总数量'] for v in all_pkgs),
        # '总合计': sum(v['sum']['总合计'] for v in all_pkgs),
        '总毛重': sum(v['sum']['总毛重'] for v in all_pkgs),
        '总净重': sum(v['sum']['总净重'] for v in all_pkgs),
        # '总件数': sum(v['sum']['总件数'] for v in all_pkgs),
    }
    return columns, packings, all_sum


def check(proforma_invoice, packing_list, air_waybill):
    if proforma_invoice is None:
        logger.warning('无发票文件')
        return
    if packing_list is None:
        logger.warning('无箱单文件')
        return
    file1, file2 = load_workbook(proforma_invoice), load_workbook(packing_list)
    sheet1, sheet2 = file1.worksheets[0], file2.worksheets[0]

    columns1, invoices, all_sum1 = read_invoice(sheet1)
    columns2, packings, all_sum2 = read_packing(sheet2)
    # 发票 VS 箱单
    keys1 = set(invoices.keys()) # dict_keys(set-like object) -> set
    keys2 = set(('00' + k) for k in packings.keys())
    if keys1 == keys2:
        for key in keys1:
            v1, v2 = invoices[key], packings[re00.sub('', key)]
            # NOTE: 只核对总数量，总净重（发票文件无该信息），总毛重，总件数不核对
            checker.check_1_2_qty(logger, key, v1, v2)
            # 校验箱单总毛重大于总净重
            checker.check_2_net_gross_weight(logger, key, v2)
    else:
        checker.write_12_inovice_diff(logger, invoices, keys1 - keys2, '发票号不在箱单中')
        checker.write_12_inovice_diff(logger, packings, (re00.sub('', key) for key in (keys2 - keys1)), '发票号不在发票中')

    logger.info('文件交互核对完成，开始序列化')
    checker.write_errors(logger, (sheet1, invoices, columns1), (sheet2, packings, columns2))

    file1.save(proforma_invoice)
    file2.save(packing_list)

