# -*- coding: utf-8 -*-
__author__ = 'newdefence@163.com'
__date__ = '2022/08/11 15:23' 

# import sys
# reload(sys)
# sys.setdefaultencoding('utf8')

import logging
from openpyxl import load_workbook, Workbook

logger = logging.getLogger()

# TODO: excel 文件需要统一表头(第二行表头文字一定要一直，否则没法处理，或者需要根据一个文件一种处理方式)
def read_invoice(workbook):
    sheet = workbook.worksheets[0]
    # print('min_row: %s' % sheet.min_row)
    # print('min_column: %s' % sheet.min_column)
    # print('max_row: %s' % sheet.max_row)
    # print('max_column: %s' % sheet.max_column)
    max_row, max_column = sheet.max_row, sheet.max_column
    # 发票号，PO号，物料号，数量，单价，合计，总数量，总合计，总件数，总毛重，总净重

def read_packing(workbook):
    # 发票号，PO号，物料号，数量，单价，合计，总数量，总合计，总件数，总毛重，总净重
    pass

def check(proforma_invoice, packing_list, air_warbill):
    if proforma_invoice is None:
        logger.warn('无发票文件')
        return
    invoice = load_workbook(proforma_invoice, read_only = True)
    packing = load_workbook(packing_list, read_only = True) if packing_list else None
    air = load_workbook(air_warbill, read_only = True) if air_warbill else None

    # excel 文件其实没必要写三个，一个文件三个 sheet 即可
    read_invoice(invoice)
    if packing:
        read_invoice(packing)
    if air:
        read_invoice(air)

