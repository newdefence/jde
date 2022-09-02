# -*- coding: utf-8 -*-
__author__ = 'newdefence@163.com'
__date__ = '2022/09/01 23:37'

from decimal import Decimal
import logging
import re

from openpyxl import load_workbook

import checker
import reader

'''
杜邦
有提单则为空运，无提单则为物流园
杜邦空运无excel，物流园有excel
物流园
1.（所有供应商统一逻辑）发票跟箱单检验逻辑：相同的发票号单个料号总数量校验，单个发票号总数量校验，所有发票号总数量校验
箱单上总毛重大于总净重

Excel校验
发票号对应的物料号检验数量和净重



空运：
1.（所有供应商统一逻辑）发票跟箱单检验逻辑：相同的发票号单个料号总数量校验，单个发票号总数量校验，所有发票号总数量校验
箱单上总毛重大于总净重

2.箱单跟提单：箱单上的总托盘数和总箱数还有总毛重（上下差百分之三合理范围，写入识别文件）进行提单校验
3.箱单总净重小于提单上的总毛重（如果大于等于则提示异常并返回）

'''


logger = logging.getLogger()
reSM = re.compile(r'^([\d,.]+)\s+SM')  # 11,902.2 SM
reKG = re.compile(r'^([\d,.]+)\s+KG')  # 1,690.11 KG
rePLT = re.compile(r'^([\d,]+)PLT')  # 13PLT
reCTN = re.compile(r'^([\d,]+)CTN')  # 425CTN


def Decimal2(reg, origin):
    return Decimal(reg.findall(origin)[0].replace(',', ''))


def read_invoice(sheet):
    """文件自查"""
    invoices, columns = reader.read_sheet12(sheet, ('发票号', 'po号', '物料号', '空运校验点'),
                                    ('数量', '单价', '合计', '总数量', '总净重', '总毛重', '总合计', '总件数'))
    all_invoices = invoices.values()
    for invoice in all_invoices:
        # 总数量为空
        # 总净重为空
        # 总毛重为空
        # 总件数为空
        checker.check_lt0(invoice, ('数量', '单价', '合计', '总合计'))
        checker.check_invoice_value1(invoice)
    # NOTE: 没有汇总核对需求
    logger.info('发票文件核对完成')
    return columns, invoices, None


def read_packing(sheet):
    packings, columns = reader.read_sheet12(sheet, ('发票号', 'po号', '物料号', '数量', '净重', '总净重', '总毛重', '总托盘数', '总箱数'), ('毛重', '总数量', '总件数'))
    # 每一个发票号明细总数量，总净重，总毛重全部相等，总件数全部相同
    all_pkgs = packings.values()
    check_总托盘数 = checker.gen_check_total('总托盘数')
    check_总箱数 = checker.gen_check_total('总箱数')
    for invoice in all_pkgs:
        for d in invoice['details']:
            d['数量'] = Decimal2(reSM, d['数量'])  # '11,902.2 SM' -> 11902.2
            d['净重'] = Decimal2(reKG, d['净重'])  # '1,690.11 KG' -> 1690.11
            d['总净重'] = Decimal2(reKG, d['总净重'])
            d['总毛重'] = Decimal2(reKG, d['总毛重'])
            d['总托盘数'] = Decimal2(rePLT, d['总托盘数'])  # '12PLT' -> 12
            d['总箱数'] = Decimal2(reKG, d['总箱数'])  # '425CTN' -> 425
        # 毛重为空
        # 总数量为空
        # 总件数为空（使用总托盘数√|总箱数代替）
        checker.check_lt0(invoice['details'], ('数量', '净重', '总净重', '总毛重', '总托盘数', '总箱数'))
        checker.check_net_weight2(invoice)
        checker.check_gross_weight1(invoice)
        check_总托盘数(invoice)
        check_总箱数(invoice)
    logger.info('箱单文件核对完成')
    return columns, packings, None


def check(proforma_invoice, packing_list, air_waybill, excel):
    file1, file2 = load_workbook(proforma_invoice), load_workbook(packing_list)
    sheet1, sheet2 = file1.worksheets[0], file2.worksheets[0]
    file3, sheet3, airs = None, None, None
    file4, sheet4, cpns = None, None, None

    columns1, invoices, _ = read_invoice(sheet1)
    columns2, packings, _ = read_packing(sheet2)
    if air_waybill:
        # TODO: 校验规则，以及样例数据
        pass
    if excel:
        file4 = load_workbook(excel)
        sheet4 = file4.worksheets[0]
        cpns = reader.read_sheet4(sheet4, 'F')  # F列是发票号
    # 发票 VS 箱单
    keys1 = invoices.keys() # dict_keys(set-like object) -> set
    keys2 = packings.keys()
    if keys1 == keys2:
        for key in keys1:
            v1, v2 = invoices[key], packings[key]
            # NOTE: 只核对总数量，总净重（发票文件无该信息），总毛重，总件数不核对
            checker.check_1_2_qty(logger, key, v1, v2)
            # 校验箱单总毛重大于总净重
            checker.check_2_net_gross_weight(logger, key, v2)
            # TODO 数量需要逐个比对
    else:
        checker.write_12_inovice_diff(logger, invoices, keys1 - keys2, '发票号不在箱单中')
        checker.write_12_inovice_diff(logger, packings, keys2 - keys1, '发票号不在发票中')

    # TODO 箱单 VS Excel校验
    if excel:
        dict4 = {}
        for row in cpns:
            pass

    logger.info('文件交互核对完成，开始序列化')
    checker.write_errors(logger, (sheet1, invoices, columns1), (sheet2, packings, columns2))

    file1.save(proforma_invoice)
    file2.save(packing_list)

