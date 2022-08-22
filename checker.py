# -*- coding: utf-8 -*-
__author__ = 'newdefence@163.com'
__date__ = '2022/08/19 10:25'

from openpyxl.utils import get_column_letter
# def check_sum1_sum2(details, key, sum1):
#     sum2 = sum(d[key] for d in details)


def check_all_sum(sum_row, all_invoices, keys):
    """
        all_qty = sum(v['sum']['总数量'] for v in all_invoices)
        all_invoice_value = sum(v['sum']['总合计'] for v in all_invoices)
        all_gross_weight = sum(v['sum']['总毛重'] for v in all_invoices)
        if all_sum['总数量'] != all_qty:
            all_sum['errors'].add('SUM(总数量): %s %s' % (all_sum['总数量'], all_qty))
        if all_sum['总合计'] != all_invoice_value:
            all_sum['errors'].add('SUM(总合计): %s %s' % (all_sum['总合计'], all_invoice_value))
        if all_sum['总合计'] != all_gross_weight:
            all_sum['errors'].add('SUM(总合计): %s %s' % (all_sum['总合计'], all_gross_weight))
    """
    for key in keys:
        sum1 = sum(v['sum'][key] for v in all_invoices)
        if sum_row[key] != sum1:
            sum_row['errors'].add('SUM(%s): %s %s' % (key, sum_row[key], sum1))


def check_qty12(invoice):
    sum1, details = invoice['sum'], invoice['details']
    total_qty = sum1['总数量']
    total_qty1 = sum(d['数量'] for d in details)
    if total_qty != total_qty1:
        tuple(d['errors'].add('SUM(数量): %s %s' % (total_qty, total_qty1)) for d in details)
    tuple(d['errors'].add('总数量: %s %s' % (d['总数量'], total_qty)) for d in details if (d['总数量'] != total_qty))


def check_invoice_value1(invoice):
    sum1, details = invoice['sum'], invoice['details']
    total_invoice_value = sum1['总合计']
    total_invoice_value1 = sum(d['合计'] for d in details)
    tuple(d['errors'].add('合计: %s %s' % (d['合计'], d['数量'] * d['单价'])) for d in details if (d['合计'] != d['数量'] * d['单价']))
    if total_invoice_value != total_invoice_value1:
        tuple(d['errors'].add('SUM(总合计): %s %s' % (total_invoice_value, total_invoice_value1)) for d in details)
    tuple(d['errors'].add('总合计: %s %s' % (d['总合计'], total_invoice_value)) for d in details if (d['总合计'] != total_invoice_value))


def check_gross_weight1(invoice):
    total_gross_weight = invoice['sum']['总毛重']
    tuple(d['errors'].add('总毛重: %s %s' % (d['总毛重'], total_gross_weight)) for d in invoice['details'] if (d['总毛重'] != total_gross_weight))


def check_gross_weight2(invoice):
    sum1, details = invoice['sum'], invoice['details']
    total_gross_weight = sum1['总毛重']
    total_gross_weight1 = sum(d['毛重'] for d in details)
    if total_gross_weight != total_gross_weight1:
        tuple(d['errors'].add('SUM(毛重): %s %s' % (total_gross_weight, total_gross_weight1)) for d in details)
    tuple(d['errors'].add('总毛重: %s %s' % (d['总毛重'], total_gross_weight)) for d in details if (d['总毛重'] != total_gross_weight))


def check_net_weight1(invoice):
    total_net_weight = invoice['sum']['总净重']
    tuple(d['errors'].add('总净重: %s %s' % (d['总净重'], total_net_weight)) for d in invoice['details'] if (d['总净重'] != total_net_weight))


def check_net_weight2(invoice):
    sum1, details = invoice['sum'], invoice['details']
    total_net_weight = sum1['总净重']
    total_net_weight1 = sum(d['净重'] for d in details)
    if total_net_weight != total_net_weight1:
        tuple(d['errors'].add('SUM(净重): %s %s' % (total_net_weight, total_net_weight1)) for d in details)
    tuple(d['errors'].add('总净重: %s %s' % (d['总净重'], total_net_weight)) for d in details if (d['总净重'] != total_net_weight))


def check_piece12(invoice): # 总件数
    piece = invoice['sum']['总件数']
    tuple(d['errors'].add('总件数: %s %s' % (d['总件数'], piece)) for d in invoice['details'] if (d['总件数'] != piece))


def check_1_2_qty(logger, key, v1, v2): # 发票文件发票号，箱单文件发票号
    sum1, sum2 = v1['sum']['总数量'], v2['sum']['总数量']
    if sum1 == sum2:
        logger.info('发票 VS 箱单：%s %s总数量相同', key, '相同发票号单个料号' if len(v1['details']) > 1 else '单个发票号')
        return True
    else:
        msg = '发票VS提单总数量: %s %s' % (sum1, sum2)
        logger.warning('发票 VS 提单总数量错误：%s %s %s' % (key, sum1, sum2))
        tuple(d['errors'].add(msg) for d in v1['details'])
        tuple(d['errors'].add(msg) for d in v2['details'])
        return False


def check_2_net_gross_weight(logger, key, v2):
    ok = True
    for d in v2['details']:
        if d['总毛重'] <= d['总净重']:
            ok = False
            logger.warning('发票 VS 箱单：%s 总毛重:%s ≤ 总净重: %s', (key, d['总毛重'], d['总净重']))
            d['errors'].add('总毛重:%s ≤ 总净重: %s' % (d['总毛重'], d['总净重']))
        else:
            logger.info('发票 VS 箱单：%s 总毛重 ＞ 总净重', key)
    return ok


def write_12_inovice_diff(logger, host, keys, msg):
    # 发票号不在箱单/发票中
    if not keys:
        return
    for key in keys:
        logger.warning('发票 VS 箱单：%s %s', key, msg)
        tuple(d['errors'].add(msg) for d in host[key]['details'])


def _write_errors_details(sheet, details, col1, col2):
    has_error = False
    for d in details:
        row, errors, warnings = d['row'][0].row, d['errors'], d.get('warnings')
        if errors:
            has_error = True
        if warnings:
            errors = errors + warnings
        if errors:
            sheet['%s%s' % (get_column_letter(col1), row)] = None
            sheet.cell(row, col2, '，'.join(errors))
        else:
            sheet.cell(row, col1, '可入账')
            sheet['%s%s' % (get_column_letter(col2), row)] = None
    return has_error


def check_errors_all_sum(logger, sheet1, sum_row, columns):
    row_idx = sum_row['row'][0].row
    if sum_row['errors']:
        msg = '，'.join(sum_row['errors'])
        logger.warning('汇总信息错误：%s' % msg)
        sheet1['%s%s' % (get_column_letter(columns['可否入账'] + 1), row_idx)] = None
        sheet1.cell(row_idx, columns['异常信息'] + 1, msg)
        return True
    logger.info('汇总信息正常')
    sheet1.cell(row_idx, columns['可否入账'] + 1, '可入账')
    sheet1['%s%s' % (get_column_letter(columns['异常信息'] + 1), row_idx)] = None
    return False

def write_errors(logger, file1, file2, file3=None):
    errors = 0
    sheet1, host1, columns1 = file1
    col1, col2 = columns1['可否入账'] + 1, columns1['异常信息'] + 1
    for v1 in host1.values():
        errors += _write_errors_details(sheet1, v1['details'], col1, col2)
    sheet2, host2, columns2 = file2
    col1, col2 = columns2['可否入账'] + 1, columns2['异常信息'] + 1
    for v2 in host2.values():
        errors += _write_errors_details(sheet2, v2['details'], col1, col2)
    if file3:
        sheet3, details3, columns3 = file3
        col1, col2 = columns3['可否入账'] + 1, columns3['异常信息'] + 1
        errors += _write_errors_details(sheet3, details3, col1, col2)
    if errors:
        logger.warning('文件校验完成，错误已标注')
    else:
        logger.info('文件校验完成，没有错误信息')
