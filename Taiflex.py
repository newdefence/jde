# -*- coding: utf-8 -*-
__author__ = 'newdefence@163.com'
__date__ = '2022/08/16 17:43'

import logging
from decimal import Decimal

from openpyxl import load_workbook
from ordered_set import OrderedSet

import checker
import reader # import read_sheet12, write_row_errors, value


'''
台虹：（需要多识别一个出售合同）taiflex
    有提单则为空运（提单上的除了发票和箱单外还有其他的页则为空运）
    物流园
    1.（所有供应商统一逻辑）发票跟箱单检验逻辑：相同的发票号单个料号总数量校验，单个发票号总数量校验，所有发票号总数量校验
    箱单上总毛重大于总净重

    发票跟出售合同检验：各个发票号中的总数量和总合计、条款进行合同检验

空运：
1.（所有供应商统一逻辑）发票跟箱单检验逻辑：相同的发票号单个料号总数量校验，单个发票号总数量校验，所有发票号总数量校验；
箱单上总毛重大于总净重
2.箱单跟提单：箱单上的总托盘数和总箱数还有总毛重（上下差百分之三合理范围，写入识别文件）进行提单校验
3.箱单总净重小于提单上的总毛重（如果大于等于则提示异常并返回）
'''


logger = logging.getLogger()


def read_invoice(sheet):
    """文件自查"""
    # 发票号，地址，原产国，PO号，物料号，数量，单价，合计，总数量，总合计，总件数，总毛重，总净重
    # { '发票号': 1, '物料号': 3, ... }
    # { 'invoice_no': { 'sum': { '总合计': 1, '总毛重': 2, ... }, 'details': [{ 'PO号': 'xxx', '物料号': '', ... }, ...] } }
    all_sum = {}
    def fn_sum(row, columns):
        all_sum.update({
            'row': row, 'errors': OrderedSet(),
            '总数量': reader.value(row[columns['总数量']]),
            '总合计': reader.value(row[columns['总合计']]),
            '总毛重': reader.value(row[columns['总毛重']]),
            '总净重': reader.value(row[columns['总净重']]),
            # '总件数': value(row[columns['总件数']]),
            # NOTE: 总件数，总净重没有
        })

    invoices, columns = reader.read_sheet12(sheet, ('发票号', 'po号', '物料号'), ('数量', '单价', '合计', '总数量', '总合计', '总件数', '总净重', '总毛重'), fn_sum)
    all_invoices = invoices.values()
    for invoice in all_invoices:
        checker.check_qty12(invoice)
        checker.check_invoice_value1(invoice)
        checker.check_gross_weight1(invoice)
        checker.check_net_weight1(invoice) # 毛重没法比对，只能和自己比对
        checker.check_piece12(invoice)
    checker.check_all_sum(all_sum, all_invoices, ('总数量', '总合计', '总毛重'))
    # NOTE: 没有总件数和总净重的核对需求
    logger.info('发票文件核对完成')
    return columns, invoices, all_sum


def read_packing(sheet):
    # 发票号，PO号，物料号，数量，单价，合计，总数量，总合计，总件数，总毛重，总净重
    # { 'invoice_no': { 'sum': { '总合计': 1, '总毛重': 2, ... }, 'details': [{ 'PO号': 'xxx', '物料号': '', ... }, ...] } }
    packings, columns = reader.read_sheet12(sheet, ('发票号', 'po号', '物料号'), ('数量', '净重', '毛重', '总数量', '总净重', '总毛重', '总件数', '总托盘数', '总箱数'))
    # 每一个发票号明细总数量，总净重，总毛重全部相等，总件数全部相同，并等于该发票号所有数量合计
    all_pkgs, all_sum = packings.values(), {'总毛重': 0, '总净重': 0}
    for invoice in all_pkgs:
        sum1 = invoice['sum']
        all_sum['总毛重'] += sum1['总毛重']
        all_sum['总净重'] += sum1['总净重']
        # TODO CONFIRM: 总托盘数，总箱数，总件数 不累加 ?
        all_sum['总托盘数'] = sum1['总托盘数']
        all_sum['总箱数'] = sum1['总箱数']
        all_sum['总件数'] = sum1['总件数']
        checker.check_qty12(invoice)
        checker.check_net_weight2(invoice)
        checker.check_gross_weight2(invoice)
        checker.check_piece12(invoice)
    logger.info('箱单文件核对完成')
    return columns, packings, all_sum


def read_air(sheet):
    # 逻辑如何处理，只有提运单号，处理那些字段呢？
    columns = None  # { '发票号': 1, '物料号': 3, ... }
    details, columns = reader.read_sheet3(sheet, ('主运单号', '分运单号'), ('托盘数', '总毛重', '总件数'))
    all_sum = {
        '托盘数': sum(d['托盘数'] for d in details),
        '总毛重': sum(d['总毛重'] for d in details),
        '总件数': sum(d['总件数'] for d in details),
    }
    return columns, details, all_sum


def check(proforma_invoice, packing_list, air_waybill):
    if proforma_invoice is None:
        logger.warning('无发票文件')
        return
    if packing_list is None:
        logger.warning('无箱单文件')
        return
    file1, file2 = load_workbook(proforma_invoice), load_workbook(packing_list)
    sheet1, sheet2 = file1.worksheets[0], file2.worksheets[0]

    columns1, invoices, all_sum1 = read_invoice(sheet1)
    columns2, packings, all_sum2 = read_packing(sheet2)
    # 发票 VS 箱单
    keys1 = invoices.keys()
    keys2 = packings.keys()
    if keys1 == keys2:
        for key in keys1:
            # NOTE: 只核对总数量，总净重（发票文件无该信息），总毛重，总件数不核对
            checker.check_1_2_qty(logger, key, invoices[key], packings[key])
            # 校验箱单总毛重大于总净重
            checker.check_2_net_gross_weight(logger, key, packings[key])
    else:
        checker.write_12_inovice_diff(logger, invoices, keys1 - keys2, '发票号不在箱单中')
        checker.write_12_inovice_diff(logger, packings, keys2 - keys1, '发票号不在发票中')

    # 箱单 VS 提单，如果出现错误，只写提单
    if air_waybill:
        file3 = load_workbook(air_waybill)
        sheet3 = file3.worksheets[0]
        columns3, airs, all_sum3 = read_air(sheet3)
        def write_air_errors(msg):
            logger.warn(msg)
            tuple(d['errors'].add(msg) for d in airs)
        if all_sum2['总托盘数'] != all_sum3['托盘数']:
            write_air_errors('托盘数与箱单不符')
        # 目前文件没有 总箱数 此列
        if all_sum2['总箱数'] != all_sum3.get('总箱数'):
            write_air_errors('总箱数与箱单不符:该文件无`总箱数`列')
        # 总毛重误差3%内正常
        diff_gross_weight = all_sum2['总毛重'] - all_sum3['总毛重']
        if diff_gross_weight:
            if diff_gross_weight >= all_sum2['总毛重'] * Decimal(0.03):
                write_air_errors('总毛重与箱单不符：超过3%')
            else:
                tuple(d.setdefault('warnings', set()).add('总毛重与箱单误差3%以内') for d in airs)
        if all_sum2['总净重'] >= all_sum3['总毛重']:
            write_air_errors('箱单总净重需小于提单总毛重')
    
    logger.info('文件交互核对完成，开始序列化')
    checker.check_errors_all_sum(logger, sheet1, all_sum1, columns1)
    checker.write_errors(logger, (sheet1, invoices, columns1), (sheet2, packings, columns2),
        (sheet3, airs, columns3) if air_waybill else None)

    file1.save(proforma_invoice)
    file2.save(packing_list)
    if air_waybill:
        file3.save(air_waybill)
