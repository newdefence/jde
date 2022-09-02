# -*- coding: utf-8 -*-
__author__ = 'newdefence@163.com'
__date__ = '2022/09/01 08:52'

from decimal import Decimal
import logging
import re

from openpyxl import load_workbook

import checker
import reader

'''
SANSHIN：
只有空运
空运
1.（所有供应商统一逻辑）发票跟箱单检验逻辑：相同的发票号单个料号总数量校验，单个发票号总数量校验，所有发票号总数量校验；
箱单上总毛重大于总净重
2.箱单跟提单：箱单上的总托盘数和总箱数，还有总毛重（上下差百分之三合理范围，写入识别文件）进行提单校验
3.箱单总净重小于提单上的总毛重（如果大于等于则提示异常并返回）

'''


logger = logging.getLogger()
rePCS = re.compile(r'^([\d,]+)pcs\.')  # 123,000pcs.
reQty = re.compile(r'^([\d,]+)\s+CARTON\(S\)\s+ONLY\.')  # 3 CARTON(S) ONLY.


def Decimal2(reg, origin):
    return Decimal(reg.findall(origin)[0].replace(',', ''))


def read_invoice(sheet):
    """文件自查"""
    # 发票号，地址，原产国，PO号，物料号，数量，单价，合计，总数量，总合计，总件数，总毛重，总净重
    # { '发票号': 1, '物料号': 3, ... }
    # { 'invoice_no': { 'sum': { '总合计': 1, '总毛重': 2, ... }, 'details': [{ 'PO号': 'xxx', '物料号': '', ... }, ...] } }

    # NOTE: 立寰发票文件没有汇总行
    invoices, columns = reader.read_sheet12(sheet, ('发票号', 'po号', '物料号', '总数量'),
                                    ('数量', '单价', '合计', '总净重', '总毛重', '总合计', '总件数'))
    all_invoices = invoices.values()
    for invoice in all_invoices:
        for d in invoice['details']:
            d['总数量'] = Decimal2(rePCS, d['总数量'])  # '60,000 EA' -> 60000
        checker.check_qty12(invoice)
        checker.check_invoice_value1(invoice)
        # 总净重，总毛重，总件数 均为空
    # NOTE: 没有汇总核对需求
    logger.info('发票文件核对完成')
    return columns, invoices, None


def read_packing(sheet):
    packings, columns = reader.read_sheet12(sheet, ('发票号', 'po号', '物料号', '总数量', '总件数'), ('数量', '净重', '毛重', '总净重', '总毛重', '总托盘数', '总箱数'))
    # 每一个发票号明细总数量，总净重，总毛重全部相等，总件数全部相同
    all_pkgs = packings.values()
    for invoice in all_pkgs:
        for d in invoice['details']:
            d['总数量'] = Decimal2(rePCS, d['总数量'])  # '154,000pcs.' -> 154000
            d['总件数'] = Decimal2(reQty, d['总件数'])  # '42 CARTON(S) ONLY.' -> 42
        checker.check_qty12(invoice)
        checker.check_net_weight2(invoice)
        checker.check_gross_weight2(invoice)
        checker.check_piece12(invoice)
    logger.info('箱单文件核对完成')
    return columns, packings, None


def check(proforma_invoice, packing_list, air_waybill):
    if proforma_invoice is None:
        logger.warning('无发票文件')
        return
    if packing_list is None:
        logger.warning('无箱单文件')
        return
    file1, file2 = load_workbook(proforma_invoice), load_workbook(packing_list)
    sheet1, sheet2 = file1.worksheets[0], file2.worksheets[0]

    columns1, invoices, _ = read_invoice(sheet1)
    columns2, packings, _ = read_packing(sheet2)
    # 汇总箱单SUM
    checker.write_cpn_sum_sheet(file2, packings)
    # 发票 VS 箱单
    keys1 = invoices.keys() # dict_keys(set-like object) -> set
    keys2 = packings.keys()
    if keys1 == keys2:
        for key in keys1:
            v1, v2 = invoices[key], packings[key]
            # NOTE: 只核对总数量，总净重（发票文件无该信息），总毛重，总件数不核对
            checker.check_1_2_qty(logger, key, v1, v2)
            # 校验箱单总毛重大于总净重
            checker.check_2_net_gross_weight(logger, key, v2)
    else:
        checker.write_12_inovice_diff(logger, invoices, keys1 - keys2, '发票号不在箱单中')
        checker.write_12_inovice_diff(logger, packings, keys2 - keys1, '发票号不在发票中')

    logger.info('文件交互核对完成，开始序列化')
    checker.write_errors(logger, (sheet1, invoices, columns1), (sheet2, packings, columns2))

    file1.save(proforma_invoice)
    file2.save(packing_list)

